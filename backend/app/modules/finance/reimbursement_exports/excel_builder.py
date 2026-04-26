"""
Reimbursement Exports Excel Builder

Builds Excel export files using openpyxl from the template.
"""

import io
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .constants import (
    CATEGORY_MEALS_ENTERTAINMENT,
    CATEGORY_OTHER_PROJECT,
    CATEGORY_TRANSPORTATION,
    CATEGORY_VEHICLE,
    EXCEL_TEMPLATE_PATH,
    SHEET_CATEGORY,
    SHEET_DETAIL,
)


def _copy_cell_style(src, dst):
    """Copy style from src cell to dst cell."""
    if src.has_style:
        dst.font = Font(
            name=src.font.name,
            size=src.font.size,
            bold=src.font.bold,
            italic=src.font.italic,
            vertAlign=src.font.vertAlign,
            underline=src.font.underline,
            strike=src.font.strike,
            color=src.font.color,
        )
        dst.border = src.border
        dst.fill = src.fill
        dst.number_format = src.number_format
        dst.protection = src.protection
        dst.alignment = src.alignment


def _template_path() -> str:
    from pathlib import Path

    p = Path(EXCEL_TEMPLATE_PATH)
    if not p.is_absolute():
        project_root = Path(__file__).resolve().parents[5]
        p = (project_root / p).resolve()
    return str(p)


def _doc_number_range(doc_numbers: list[int]) -> str:
    """Compress consecutive doc numbers into ranges.

    Examples:
        [1,2,3,4,5] -> "1-5"
        [1,2,5,7,8,9] -> "1-2, 5, 7-9"
    """
    if not doc_numbers:
        return ""
    nums = sorted(set(doc_numbers))
    ranges: list[str] = []
    start = nums[0]
    end = nums[0]
    for n in nums[1:]:
        if n == end + 1:
            end = n
        else:
            if start == end:
                ranges.append(str(start))
            else:
                ranges.append(f"{start}-{end}")
            start = n
            end = n
    if start == end:
        ranges.append(str(start))
    else:
        ranges.append(f"{start}-{end}")
    return ", ".join(ranges)


def build_excel(
    records: list[dict],
    category_groups: dict[str, list[dict]],
    department: str,
) -> bytes:
    """Build Excel file from template and return bytes."""
    template = _template_path()
    from pathlib import Path

    if Path(template).exists():
        wb = load_workbook(template)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title=SHEET_DETAIL)
        wb.create_sheet(title=SHEET_CATEGORY)
        # Seed detail sheet header
        ws_d = wb[SHEET_DETAIL]
        headers = ["单据编号", "日期", "用途/说明", "金额", "币种", "分类", "子分类", "使用公司/部门/项目/合同", "地点由", "地点至", "备注"]
        for col_idx, hdr in enumerate(headers, start=1):
            ws_d.cell(row=1, column=col_idx, value=hdr)
        # Seed category sheet header
        ws_c = wb[SHEET_CATEGORY]
        cat_headers = ["分类", "子分类", "金额合计", "单据编号范围"]
        for col_idx, hdr in enumerate(cat_headers, start=1):
            ws_c.cell(row=1, column=col_idx, value=hdr)

    # ------------------------------------------------------------------
    # Detail sheet
    # ------------------------------------------------------------------
    ws_detail = wb[SHEET_DETAIL]
    detail_start_row = 2

    for idx, rec in enumerate(records, start=0):
        row = detail_start_row + idx
        # 用途/说明: 优先 remark, 没有则 order_name
        description = rec.get("remark") or rec["order_name"]
        # 交通费用地点由/至留空
        from_location = ""
        to_location = ""
        ws_detail.cell(row=row, column=1, value=rec["document_number"])
        ws_detail.cell(row=row, column=2, value=rec["purchase_date"])
        ws_detail.cell(row=row, column=3, value=description)
        ws_detail.cell(row=row, column=4, value=float(rec["amount"]))
        ws_detail.cell(row=row, column=5, value=rec["currency"])
        ws_detail.cell(row=row, column=6, value=rec["category"])
        ws_detail.cell(row=row, column=7, value=rec.get("subcategory") or "")
        ws_detail.cell(row=row, column=8, value=department)
        ws_detail.cell(row=row, column=9, value=from_location)
        ws_detail.cell(row=row, column=10, value=to_location)
        ws_detail.cell(row=row, column=11, value=rec.get("remark") or "")

    # ------------------------------------------------------------------
    # Category sheet
    # ------------------------------------------------------------------
    ws_cat = wb[SHEET_CATEGORY]
    cat_start_row = 2
    current_row = cat_start_row

    categories = [
        CATEGORY_TRANSPORTATION,
        CATEGORY_MEALS_ENTERTAINMENT,
        CATEGORY_VEHICLE,
        CATEGORY_OTHER_PROJECT,
    ]

    for cat in categories:
        cat_records = category_groups.get(cat, [])
        if not cat_records:
            continue

        if cat == CATEGORY_OTHER_PROJECT:
            # Group by subcategory
            sub_groups: dict[str, list[dict]] = {}
            for crec in cat_records:
                sub = crec.get("subcategory") or "未分类"
                sub_groups.setdefault(sub, []).append(crec)
            for sub, sub_recs in sub_groups.items():
                total = sum(float(r["amount"]) for r in sub_recs)
                doc_nums = [r["document_number"] for r in sub_recs]
                ws_cat.cell(row=current_row, column=1, value=cat)
                ws_cat.cell(row=current_row, column=2, value=sub)
                ws_cat.cell(row=current_row, column=3, value=total)
                ws_cat.cell(row=current_row, column=4, value=_doc_number_range(doc_nums))
                current_row += 1
        else:
            total = sum(float(r["amount"]) for r in cat_records)
            doc_nums = [r["document_number"] for r in cat_records]
            ws_cat.cell(row=current_row, column=1, value=cat)
            ws_cat.cell(row=current_row, column=2, value="")
            ws_cat.cell(row=current_row, column=3, value=total)
            ws_cat.cell(row=current_row, column=4, value=_doc_number_range(doc_nums))
            current_row += 1

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
