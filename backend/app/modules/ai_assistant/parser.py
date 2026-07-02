"""In-memory document parsers for the AI assistant.

Files are read directly from the request stream into memory, parsed for text,
and then discarded. Nothing is written to disk.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Sequence

from fastapi import UploadFile


class FileParseError(ValueError):
    """Raised when a file cannot be parsed or has an unsupported format."""


class FileContentTooLargeError(ValueError):
    """Raised when uploaded files exceed size or parsed-text limits."""


MAX_FILE_SIZE_MB = 10
MAX_TOTAL_SIZE_MB = 50
MAX_TOTAL_CHARS = 80_000

_ALLOWED_EXTENSIONS = frozenset({".pdf", ".docx", ".txt", ".md", ".csv", ".xlsx"})


def _check_extension(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in _ALLOWED_EXTENSIONS:
        raise FileParseError(
            f"不支持的文件格式：{filename}。允许：{', '.join(sorted(_ALLOWED_EXTENSIONS))}"
        )
    return suffix


def _decode_text(content: bytes) -> str:
    """Try common encodings; fall back to latin-1 so we never crash on bytes."""
    for encoding in ("utf-8", "gbk", "gb2312", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", errors="replace")


def _parse_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover
        raise FileParseError("PDF 解析依赖未安装") from exc

    reader = PdfReader(io.BytesIO(content))
    texts: list[str] = []
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            texts.append(page_text)
    return "\n".join(texts)


def _parse_docx(content: bytes) -> str:
    try:
        from docx import Document
    except ImportError as exc:  # pragma: no cover
        raise FileParseError("Word 解析依赖未安装") from exc

    document = Document(io.BytesIO(content))
    return "\n".join(p.text for p in document.paragraphs)


def _parse_xlsx(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise FileParseError("Excel 解析依赖未安装") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets: list[str] = []
    for sheet in wb.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append("\t".join("" if value is None else str(value) for value in row))
        if rows:
            sheets.append(f"--- Sheet: {sheet.title} ---\n" + "\n".join(rows))
    return "\n\n".join(sheets)


def _parse_text(content: bytes) -> str:
    return _decode_text(content)


def _extract_text_from_bytes(filename: str, content: bytes) -> str:
    suffix = _check_extension(filename)
    try:
        if suffix == ".pdf":
            return _parse_pdf(content)
        if suffix == ".docx":
            return _parse_docx(content)
        if suffix == ".xlsx":
            return _parse_xlsx(content)
        return _parse_text(content)
    except FileParseError:
        raise
    except Exception as exc:
        raise FileParseError(f"解析文件 {filename} 失败：{exc}") from exc


def extract_text(file: UploadFile) -> str:
    """Extract text from a single uploaded file."""
    filename = file.filename or "unknown"
    content = file.file.read()
    size_mb = len(content) / (1024 * 1024)
    if size_mb > MAX_FILE_SIZE_MB:
        raise FileParseError(
            f"文件 {filename} 大小为 {size_mb:.1f}MB，超过 {MAX_FILE_SIZE_MB}MB 限制"
        )
    return _extract_text_from_bytes(filename, content)


def parse_upload_files(files: Sequence[UploadFile]) -> str:
    """Parse multiple uploaded files and return a single concatenated context string."""
    if not files:
        raise FileParseError("未上传文件")

    contents: list[tuple[str, bytes]] = []
    total_size = 0
    for file in files:
        filename = file.filename or "unknown"
        content = file.file.read()
        total_size += len(content)
        size_mb = total_size / (1024 * 1024)
        if size_mb > MAX_TOTAL_SIZE_MB:
            raise FileContentTooLargeError(
                f"批量文件总大小超过 {MAX_TOTAL_SIZE_MB}MB 限制"
            )
        contents.append((filename, content))

    parts: list[str] = []
    for filename, content in contents:
        text = _extract_text_from_bytes(filename, content)
        if not text.strip():
            text = "（该文件未解析出文本内容）"
        parts.append(f"===== {filename} =====\n{text.strip()}")

    combined = "\n\n".join(parts)
    if len(combined) > MAX_TOTAL_CHARS:
        raise FileContentTooLargeError(
            f"文件解析后总文本超过 {MAX_TOTAL_CHARS} 字符，请精简内容后重试"
        )
    return combined
